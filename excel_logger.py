"""
Módulo de Logging en Excel para Bot de Trading

Registra cada análisis completo en trading_history.xlsx con:
- Datos de todos los 5 pasos (Regresión, Expertos, IA, Veredicto)
- Seguimiento de resultados reales
- Cálculo de precisión de cada componente
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime
import os
from pathlib import Path

# Ruta del archivo Excel
EXCEL_FILE = Path(__file__).parent / "trading_history.xlsx"

# Definición de columnas
COLUMNS = [
    # Identificación
    'Timestamp', 'Símbolo', 'Tipo_Activo',
    
    # Paso 2: Regresión Lineal
    'Reg_Tendencia', 'Reg_Confianza_R2', 'Reg_Recomendación',
    
    # Paso 3: Expertos Técnicos
    'RSI_Valor', 'RSI_Señal',
    'MACD_Valor', 'MACD_Señal',
    'Heikin_Ashi', 'EMA_Cross', 'Bollinger',
    'Panel_Veredicto', 'Panel_Confianza',
    
    # Paso 4: IA
    'IA_Señal', 'IA_Confianza', 'IA_Razón',
    
    # Paso 5: Veredicto Final
    'Veredicto_Final', 'Razón_Final', 'Acción_Ejecutada',
    
    # Seguimiento (se actualiza después)
    'Precio_Entrada', 'SL', 'TP',
    'Precio_Cierre', 'Resultado', 'Ganancia_Pips', 'Ganancia_USD',
    
    # Validación retrospectiva
    'Reg_Acertó', 'Panel_Acertó', 'IA_Acertó'
]

def initialize_excel():
    """
    Crea el archivo Excel si no existe e inicializa las hojas con encabezados.
    """
    if EXCEL_FILE.exists():
        return  # Ya existe
    
    # Crear workbook
    wb = openpyxl.Workbook()
    
    # Eliminar hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Crear hoja "Resumen"
    ws = wb.create_sheet("Resumen", 0)
    
    # Escribir encabezados con formato
    for col_num, column_title in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajustar ancho de columnas
    for col_num in range(1, len(COLUMNS) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15
    
    # Hacer columnas de texto más anchas
    ws.column_dimensions['A'].width = 20  # Timestamp
    ws.column_dimensions['B'].width = 25  # Símbolo
    ws.column_dimensions['R'].width = 60  # IA_Razón
    ws.column_dimensions['T'].width = 40  # Razón_Final
    
    # Guardar
    wb.save(EXCEL_FILE)
    print(f"✅ Archivo Excel creado: {EXCEL_FILE}")

def log_analysis(data):
    """
    Registra un análisis completo en el Excel.
    
    Args:
        data (dict): Diccionario con todos los datos del análisis
                     Debe contener las claves correspondientes a COLUMNS
    """
    try:
        # Asegurar que el archivo existe
        initialize_excel()
        
        # Abrir workbook
        wb = openpyxl.load_workbook(EXCEL_FILE)
        
        # Hoja Resumen
        ws_resumen = wb["Resumen"]
        
        # Obtener nombre de hoja para el activo (nombre limpio)
        symbol = data.get('symbol', 'Unknown')
        sheet_name = symbol.replace(' ', '_').replace('Index', '').strip()[:31]  # Max 31 chars
        
        # Crear hoja del activo si no existe
        if sheet_name not in wb.sheetnames:
            ws_activo = wb.create_sheet(sheet_name)
            # Copiar encabezados
            for col_num, column_title in enumerate(COLUMNS, 1):
                cell = ws_activo.cell(row=1, column=col_num)
                cell.value = column_title
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws_activo = wb[sheet_name]
        
        # Preparar datos para insertar
        row_data = [
            # Identificación
            data.get('timestamp', datetime.datetime.now()),
            data.get('symbol', ''),
            data.get('tipo_activo', ''),
            
            # Paso 2: Regresión
            data.get('reg_tendencia', ''),
            data.get('reg_confianza', ''),
            data.get('reg_recomendacion', ''),
            
            # Paso 3: Expertos
            data.get('rsi_valor', ''),
            data.get('rsi_senal', ''),
            data.get('macd_valor', ''),
            data.get('macd_senal', ''),
            data.get('heikin_ashi', ''),
            data.get('ema_cross', ''),
            data.get('bollinger', ''),
            data.get('panel_veredicto', ''),
            data.get('panel_confianza', ''),
            
            # Paso 4: IA
            data.get('ia_senal', ''),
            data.get('ia_confianza', ''),
            data.get('ia_razon', ''),
            
            # Paso 5: Veredicto
            data.get('veredicto_final', ''),
            data.get('razon_final', ''),
            data.get('accion_ejecutada', ''),
            
            # Seguimiento (vacío por ahora)
            data.get('precio_entrada', ''),
            data.get('sl', ''),
            data.get('tp', ''),
            '', '', '', '',  # Precio_Cierre, Resultado, Ganancia_Pips, Ganancia_USD
            
            # Validación (se calcula después)
            '', '', ''  # Reg_Acertó, Panel_Acertó, IA_Acertó
        ]
        
        # Insertar en hoja Resumen
        ws_resumen.append(row_data)
        
        # Insertar en hoja del activo
        ws_activo.append(row_data)
        
        # Aplicar formato a la nueva fila
        last_row = ws_resumen.max_row
        _apply_row_formatting(ws_resumen, last_row, data)
        
        last_row_activo = ws_activo.max_row
        _apply_row_formatting(ws_activo, last_row_activo, data)
        
        # Guardar
        wb.save(EXCEL_FILE)
        print(f"📊 Análisis registrado en Excel: {symbol}")
        
    except Exception as e:
        print(f"❌ Error al registrar en Excel: {str(e)}")
        import traceback
        traceback.print_exc()

def _apply_row_formatting(ws, row_num, data):
    """
    Aplica formato condicional a una fila según los datos.
    """
    # Colorear señales
    accion = data.get('accion_ejecutada', '')
    
    # Columna de Acción Ejecutada (col 21)
    cell = ws.cell(row=row_num, column=21)
    if accion == 'COMPRA':
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell.font = Font(color="006100", bold=True)
    elif accion == 'VENTA':
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.font = Font(color="9C0006", bold=True)
    elif accion == 'ESPERAR':
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        cell.font = Font(color="9C6500")
    
    # Formato de fecha en columna A
    ws.cell(row=row_num, column=1).number_format = 'YYYY-MM-DD HH:MM:SS'

def update_result(symbol, timestamp, precio_cierre, ganancia_usd, resultado):
    """
    Actualiza el resultado real de una operación en el Excel.
    
    Args:
        symbol (str): Símbolo del activo
        timestamp (datetime): Fecha/hora del análisis original
        precio_cierre (float): Precio de cierre de la operación
        ganancia_usd (float): Ganancia/pérdida en USD
        resultado (str): 'GANANCIA' o 'PÉRDIDA'
    """
    try:
        if not EXCEL_FILE.exists():
            return
        
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Resumen"]
        
        # Buscar la fila correspondiente
        for row in range(2, ws.max_row + 1):
            cell_symbol = ws.cell(row=row, column=2).value
            cell_timestamp = ws.cell(row=row, column=1).value
            
            if cell_symbol == symbol and cell_timestamp == timestamp:
                # Actualizar datos de resultado
                ws.cell(row=row, column=25).value = precio_cierre  # Precio_Cierre
                ws.cell(row=row, column=26).value = resultado       # Resultado
                ws.cell(row=row, column=28).value = ganancia_usd    # Ganancia_USD
                
                # Calcular qué paso tuvo razón
                precio_entrada = ws.cell(row=row, column=22).value
                if precio_entrada:
                    _calculate_accuracy(ws, row, precio_entrada, precio_cierre)
                
                # Guardar
                wb.save(EXCEL_FILE)
                print(f"📊 Resultado actualizado para {symbol}")
                break
                
    except Exception as e:
        print(f"❌ Error al actualizar resultado: {str(e)}")

def _calculate_accuracy(ws, row, precio_entrada, precio_cierre):
    """
    Calcula qué paso tuvo razón comparando predicción con resultado.
    """
    # Obtener predicciones
    reg_rec = ws.cell(row=row, column=6).value      # Reg_Recomendación
    panel_ver = ws.cell(row=row, column=14).value   # Panel_Veredicto
    ia_senal = ws.cell(row=row, column=16).value    # IA_Señal
    
    # Determinar dirección real
    direccion_real = "COMPRA" if precio_cierre > precio_entrada else "VENTA"
    
    # Validar aciertos
    ws.cell(row=row, column=29).value = "SÍ" if reg_rec == direccion_real else "NO"     # Reg_Acertó
    ws.cell(row=row, column=30).value = "SÍ" if panel_ver in ["ALCISTA", "BAJISTA"] and \
                                               ((panel_ver == "ALCISTA" and direccion_real == "COMPRA") or \
                                                (panel_ver == "BAJISTA" and direccion_real == "VENTA")) else "NO"  # Panel_Acertó
    ws.cell(row=row, column=31).value = "SÍ" if ia_senal == direccion_real else "NO"   # IA_Acertó

def get_statistics(symbol=None):
    """
    Obtiene estadísticas de precisión de cada componente.
    
    Args:
        symbol (str, optional): Si se especifica, estadísticas solo de ese activo
    
    Returns:
        dict: Diccionario con estadísticas de precisión
    """
    try:
        if not EXCEL_FILE.exists():
            return {}
        
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Resumen"]
        
        stats = {
            'total_analisis': 0,
            'con_resultado': 0,
            'reg_aciertos': 0,
            'panel_aciertos': 0,
            'ia_aciertos': 0
        }
        
        for row in range(2, ws.max_row + 1):
            cell_symbol = ws.cell(row=row, column=2).value
            
            # Filtrar por símbolo si se especifica
            if symbol and cell_symbol != symbol:
                continue
            
            stats['total_analisis'] += 1
            
            # Si tiene resultado
            if ws.cell(row=row, column=26).value:  # Columna Resultado
                stats['con_resultado'] += 1
                
                # Contar aciertos
                if ws.cell(row=row, column=29).value == "SÍ":
                    stats['reg_aciertos'] += 1
                if ws.cell(row=row, column=30).value == "SÍ":
                    stats['panel_aciertos'] += 1
                if ws.cell(row=row, column=31).value == "SÍ":
                    stats['ia_aciertos'] += 1
        
        # Calcular porcentajes
        if stats['con_resultado'] > 0:
            stats['reg_precision'] = round((stats['reg_aciertos'] / stats['con_resultado']) * 100, 2)
            stats['panel_precision'] = round((stats['panel_aciertos'] / stats['con_resultado']) * 100, 2)
            stats['ia_precision'] = round((stats['ia_aciertos'] / stats['con_resultado']) * 100, 2)
        
        return stats
        
    except Exception as e:
        print(f"❌ Error al obtener estadísticas: {str(e)}")
        return {}

# Inicializar automáticamente al importar el módulo
initialize_excel()
